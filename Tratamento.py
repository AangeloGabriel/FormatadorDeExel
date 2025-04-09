# Desenvonvido por Ângelo Gabriel email para suporte: alvesangelo402@gmail.com
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import openpyxl
from tkinter import ttk, filedialog, messagebox
import tkinter as tk
import csv
import pandas as pd
import os
import tempfile
import openpyxl.workbook
from xlsxwriter.workbook import Workbook

# Definição de caminhos
caminho_arquivo_de_para = r"S:\S&OP e Produtos\Compartilhado\S&OP\Relatórios\Sell out Individual\BASE DE-PARA CLIENTES SKU v7.xlsx"

# Carrega a planilha de referência
def tratamento_havan_total(arquivo):
    global wb_principal
    wb_base = load_workbook(caminho_arquivo_de_para)
    ws_base = wb_base.active

    wb_principal = load_workbook(arquivo)
    ws_formatado = wb_principal.active

    # Remove quebras de texto e mesclagens
    for row in ws_formatado.iter_rows():
        for cell in row:
            if cell.alignment:
                cell.alignment = Alignment(wrap_text=False, horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical)
    
    for merge_range in list(ws_formatado.merged_cells.ranges):
        ws_formatado.unmerge_cells(str(merge_range))
    ws_formatado.freeze_panes = None
    
    # Preenchimento de células vazias com valores superiores
    for col in range(1, 6):  
        for row in range(8, ws_formatado.max_row + 1):
            cell = ws_formatado.cell(row=row, column=col)
            cell_acima = ws_formatado.cell(row=row-1, column=col)
            if cell.value is None or cell.value == "":
                cell.value = cell_acima.value
    
    # Limpeza e organização de colunas
    ws_formatado.delete_rows(1, 4)
    ws_formatado.insert_cols(7, 2)
    colunas_a_remover = [4, 8, 8, 9, 9, 9, 10, 10, 11, 11, 11]
    
    for col in colunas_a_remover:
        ws_formatado.delete_cols(col)
    ws_formatado['F3'], ws_formatado['G3'] = 'SKU GAMA', 'DESC GAMA'
    ws_formatado['L3'], ws_formatado['M3'] = 'Estoque Total', 'Venda Total'
    
    # Cálculo do estoque e vendas totais
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"L{linha}"] = (ws_formatado[f'H{linha}'].value or 0) + (ws_formatado[f'J{linha}'].value or 0)
        ws_formatado[f"M{linha}"] = (ws_formatado[f'I{linha}'].value or 0) + (ws_formatado[f'K{linha}'].value or 0)
    
    # Organiza nomes dos produtos
    ws_formatado.insert_cols(6)
    for linha in range(4, ws_formatado.max_row + 1):
        v1 = ws_formatado[f"D{linha}"].value or ""
        v2 = ws_formatado[f"E{linha}"].value or ""
        ws_formatado[f"F{linha}"] = f"{v1} - {v2}" if v2 and v2 not in v1 else v1
    
    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(4)

    # Criando mapeamento de valores
    def criar_mapa(coluna_chave, coluna_valor):
        mapa = {}
        for linha in range(4, ws_base.max_row + 1):
            chave = ws_base[f"{coluna_chave}{linha}"].value
            valor = ws_base[f"{coluna_valor}{linha}"].value
            if chave:
                mapa[chave] = valor
        return mapa
    mapa_1 = criar_mapa("K", "C")
    mapa_2 = criar_mapa("C", "D")
    
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"E{linha}"] = mapa_1.get(ws_formatado[f"D{linha}"].value, "Não encontrado")
        ws_formatado[f"F{linha}"] = mapa_2.get(ws_formatado[f"E{linha}"].value, "Não encontrado")

    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)


    def copiar_sem_coluna(ws_origem, nome_aba, coluna_excluir):
        ws_novo = wb_principal.create_sheet(title=nome_aba)
        for row in ws_origem.iter_rows():
            nova_linha = []
            for i, cell in enumerate(row, start=1):
                if i != coluna_excluir:  # Ignora a coluna desejada
                    nova_celula = ws_novo.cell(row=cell.row, column=len(nova_linha) + 1, value=cell.value)
                    nova_linha.append(cell.value)

    # Criar abas sem coluna G e M
    copiar_sem_coluna(ws_formatado, "Venda", 6)
    copiar_sem_coluna(ws_formatado, "Estoque", 7)
    primeira_aba = wb_principal.sheetnames[0]  # Nome da primeira aba
    ws_formatado = wb_principal[primeira_aba]  # Acessa a aba
    wb_principal.remove(ws_formatado)  # Remove a aba

def tratamento_havan_parcial(arquivo):
    global wb_principal
    wb_base = load_workbook(caminho_arquivo_de_para)
    ws_base = wb_base.active

    wb_principal = load_workbook(arquivo)
    ws_formatado = wb_principal.active

    # Remove quebras de texto e mesclagens
    for row in ws_formatado.iter_rows():
        for cell in row:
            if cell.alignment:
                cell.alignment = Alignment(wrap_text=False, horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical)
    
    for merge_range in list(ws_formatado.merged_cells.ranges):
        ws_formatado.unmerge_cells(str(merge_range))
    ws_formatado.freeze_panes = None
    
    # Preenchimento de células vazias com valores superiores
    for col in range(1, 6):  
        for row in range(8, ws_formatado.max_row + 1):
            cell = ws_formatado.cell(row=row, column=col)
            cell_acima = ws_formatado.cell(row=row-1, column=col)
            if cell.value is None or cell.value == "":
                cell.value = cell_acima.value
    
    # Limpeza e organização de colunas
    ws_formatado.delete_rows(1, 4)
    ws_formatado.insert_cols(7, 2)
    colunas_a_remover = [10,11,11,12,13,13]
    

    for col in colunas_a_remover:
        ws_formatado.delete_cols(col)
    ws_formatado['G3'], ws_formatado['H3'] = 'SKU GAMA', 'DESC GAMA'
    ws_formatado['M3'], ws_formatado['N3'] = 'Estoque Total', 'Venda Total'
    
    # Cálculo do estoque e vendas totais
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"M{linha}"] = (ws_formatado[f'J{linha}'].value or 0) + (ws_formatado[f'L{linha}'].value or 0)
        ws_formatado[f"N{linha}"] = (ws_formatado[f'I{linha}'].value or 0) + (ws_formatado[f'K{linha}'].value or 0)
    
    # Organiza nomes dos produtos
    ws_formatado.insert_cols(7)
    for linha in range(4, ws_formatado.max_row + 1):
        v1 = ws_formatado[f"E{linha}"].value or ""
        v2 = ws_formatado[f"F{linha}"].value or ""
        ws_formatado[f"G{linha}"] = f"{v1} - {v2}" if v2 and v2 not in v1 else v1
    
    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(4)

    # Criando mapeamento de valores
    def criar_mapa(coluna_chave, coluna_valor):
        mapa = {}
        for linha in range(4, ws_base.max_row + 1):
            chave = ws_base[f"{coluna_chave}{linha}"].value
            valor = ws_base[f"{coluna_valor}{linha}"].value
            if chave:
                mapa[chave] = valor
        return mapa
    mapa_1 = criar_mapa("K", "C")
    mapa_2 = criar_mapa("C", "D")
    
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"E{linha}"] = mapa_1.get(ws_formatado[f"D{linha}"].value, "Não encontrado")
        ws_formatado[f"F{linha}"] = mapa_2.get(ws_formatado[f"E{linha}"].value, "Não encontrado")

    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)


    def copiar_sem_coluna(ws_origem, nome_aba, coluna_excluir):
        ws_novo = wb_principal.create_sheet(title=nome_aba)
        for row in ws_origem.iter_rows():
            nova_linha = []
            for i, cell in enumerate(row, start=1):
                if i != coluna_excluir:  # Ignora a coluna desejada
                    nova_celula = ws_novo.cell(row=cell.row, column=len(nova_linha) + 1, value=cell.value)
                    nova_linha.append(cell.value)

    # Criar abas sem coluna G e M
    copiar_sem_coluna(ws_formatado, "Venda", 6)
    copiar_sem_coluna(ws_formatado, "Estoque", 7)
    primeira_aba = wb_principal.sheetnames[0]  # Nome da primeira aba
    ws_formatado = wb_principal[primeira_aba]  # Acessa a aba
    wb_principal.remove(ws_formatado)  # Remove a aba

def tratamento_lasa_site(arquivo):
    global wb_principal
    df = pd.read_csv(arquivo, delimiter=";", encoding='utf-8')
    # Converte o CSV para um Excel temporário
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        caminho_excel = tmp.name
        df.to_excel(caminho_excel, index=False)
    
    # Carrega o arquivo principal onde a nova aba será inserida
    wb_principal = load_workbook(caminho_excel)
    
    # Carrega a planilha temporária gerada a partir do CSV
    wb_temp = load_workbook(caminho_excel)
    ws_temp = wb_temp.active
    # Cria nova aba no arquivo principal
    ws_nova = wb_principal.create_sheet(title="Formatado")
    # Processa os dados da planilha temporária
    for row in ws_temp.iter_rows(values_only=True):
        primeira_coluna = row[0]
        if isinstance(primeira_coluna, str):
            separados = primeira_coluna.split(';')
        else:
            separados = [primeira_coluna]
    
        nova_linha = separados + list(row[1:])
        ws_nova.append(nova_linha)
    
    del wb_principal['Sheet1']

    ws_formatado = wb_principal["Formatado"]

    # Lista de colunas a remover — importante remover de trás pra frente!
    colunas_a_remover = [19,18,15,14,13,12,9,8,7,6,5,4,2]

    for col in colunas_a_remover:
        ws_formatado.delete_cols(col)

    ws_formatado.insert_cols(5, 2)

    ws_formatado.cell(row=1, column=5).value = "SKU GAMA"
    ws_formatado.cell(row=1, column=6).value = "DESC GAMA"

    wb_base = load_workbook(caminho_arquivo_de_para)
    ws_base = wb_base.active
    # # Criando mapeamento de valores
    def criar_mapa(coluna_chave, coluna_valor):
        mapa = {}
        for linha in range(2, ws_base.max_row + 1):
            chave = ws_base[f"{coluna_chave}{linha}"].value
            valor = ws_base[f"{coluna_valor}{linha}"].value
            if chave:
                chave = str(chave).strip()
                mapa[chave] = valor
        return mapa
    mapa_1 = criar_mapa("B", "C")
    mapa_2 = criar_mapa("B", "D")
    
    print(mapa_1)
    # print(mapa_2)

    for linha in range(2, ws_formatado.max_row + 1):
        ean = str(ws_formatado[f"D{linha}"].value).strip()
        ws_formatado[f"E{linha}"] = mapa_1.get(ean, "Não encontrado")
        ws_formatado[f"F{linha}"] = mapa_2.get(ean, "Não encontrado") 

    def copiar_sem_coluna(ws_origem, nome_aba, coluna_excluir):
        ws_novo = wb_principal.create_sheet(title=nome_aba)
        for row in ws_origem.iter_rows():
            nova_linha = []
            for i, cell in enumerate(row, start=1):
                if i != coluna_excluir:  # Ignora a coluna desejada
                    nova_celula = ws_novo.cell(row=cell.row, column=len(nova_linha) + 1, value=cell.value)
                    nova_linha.append(cell.value)

    # Criar abas sem coluna G e M
    copiar_sem_coluna(ws_formatado, "Venda", 8)
    copiar_sem_coluna(ws_formatado, "Estoque", 7)
    primeira_aba = wb_principal.sheetnames[0]  # Nome da primeira aba
    ws_formatado = wb_principal[primeira_aba]  # Acessa a aba
    wb_principal.remove(ws_formatado)  # Remove a aba

############## FUNCAO DO EXECUTAVEL ################

def selecionar_arquivo():
    global arquivo
    arquivo = filedialog.askopenfilename(title="Escolha um arquivo")
    nome_arquivo = os.path.basename(arquivo)
    label.config(text=f"Arquivo: {nome_arquivo}")

def salvar_arquivo():
    global wb_principal  # Garante que a planilha tratada esteja acessível
    caminho_escolhido = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                     filetypes=[("Excel files", "*.xlsx")])
    if caminho_escolhido:  # Só salva se o usuário escolher um caminho
        try:
            wb_principal.save(caminho_escolhido)
            messagebox.showinfo("Sucesso", "Arquivo Salvo com sucesso")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível salvar o arquivo: {e}")



def selecionar():
    global arquivo 
    opcao = combo.get()
    if opcao is None or opcao == "":
        label_resultado.config(text=f"Escolha um arquivo")
    else:
        label_resultado.config(text=f"Tratado com sucesso!!")

    if opcao is None or opcao == "":
        label_resultado.config(text="Escolha uma opção")
    elif opcao == "Havan_Total":
        tratamento_havan_total(arquivo)
    elif opcao == "Havan_Parcial":
        tratamento_havan_parcial(arquivo)
    elif opcao == "Lasa":
        tratamento_lasa_site(arquivo)
    

############### TELA DO EXECUTAVEL ################
root = tk.Tk()
root.title("Formatador de arquivos")
root.geometry("400x400")

botao_select = tk.Button(root, text="📂 Procurar arquivo", command=selecionar_arquivo)
botao_select.pack(pady=10, ipadx=10, ipady=5)

label = tk.Label(root, text="Nenhum arquivo selecionado", font=("Arial", 10))
label.pack(pady=10)

opcoes = ['Havan_Total', 'Havan_Parcial', 'Lasa']
combo = ttk.Combobox(root, values=opcoes, font=("Arial", 10))
combo.pack(pady=10)

botao_tratar = tk.Button(root, text="⚙️ Tratar", command=selecionar, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"))
botao_tratar.pack(pady=10, ipadx=10, ipady=5)

label_resultado = tk.Label(root, text="", font=("Arial", 10, "italic"))
label_resultado.pack(pady=5)

botao_tratar = tk.Button(root, text="💾 Salvar", command=salvar_arquivo, bg="#2196F3", fg="white", font=("Arial", 10, "bold"))
botao_tratar.pack(pady=10, ipadx=10, ipady=5)

root.mainloop()
