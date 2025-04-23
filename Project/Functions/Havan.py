from openpyxl import load_workbook
from openpyxl.styles import Alignment
from dotenv import load_dotenv
import os 
from pathlib import Path 
from Functions.Funcs import salvar_arquivo 
import tempfile


env_path = Path(__file__).resolve().parents[1] / 'Resources' / '.env'

# Carrega as variáveis do .env
load_dotenv(dotenv_path=env_path)

base = os.getenv("CLIENTE_BASE")


def tratamento_havan_total(arquivo):
    global wb_principal
    wb_base = load_workbook(base)
    ws_base = wb_base.active

    wb_principal = load_workbook(arquivo)
    ws_formatado = wb_principal.active

    # Remove quebras de texto e mesclagens
    for row in ws_formatado.iter_rows():
        for cell in row:
            if cell.alignment:
                cell.alignment = Alignment(wrap_text=False, horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical)
    
    for merge_range in list(ws_formatado.merged_cells.ranges):
        ws_formatado.unmerge_cells(str(merge_range))
    ws_formatado.freeze_panes = None
    
    # Preenchimento de células vazias com valores superiores
    for col in range(1, 6):  
        for row in range(8, ws_formatado.max_row + 1):
            cell = ws_formatado.cell(row=row, column=col)
            cell_acima = ws_formatado.cell(row=row-1, column=col)
            if cell.value is None or cell.value == "":
                cell.value = cell_acima.value
    
    # Limpeza e organização de colunas
    ws_formatado.delete_rows(1, 4)
    ws_formatado.insert_cols(7, 2)
    colunas_a_remover = [4, 8, 8, 9, 9, 9, 10, 10, 11, 11, 11]
    
    for col in colunas_a_remover:
        ws_formatado.delete_cols(col)
    ws_formatado['F3'], ws_formatado['G3'] = 'SKU GAMA', 'DESC GAMA'
    ws_formatado['L3'], ws_formatado['M3'] = 'Estoque Total', 'Venda Total'
    
    # Cálculo do estoque e vendas totais
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"L{linha}"] = (ws_formatado[f'H{linha}'].value or 0) + (ws_formatado[f'J{linha}'].value or 0)
        ws_formatado[f"M{linha}"] = (ws_formatado[f'I{linha}'].value or 0) + (ws_formatado[f'K{linha}'].value or 0)
    
    # Organiza nomes dos produtos
    ws_formatado.insert_cols(6)
    for linha in range(4, ws_formatado.max_row + 1):
        v1 = ws_formatado[f"D{linha}"].value or ""
        v2 = ws_formatado[f"E{linha}"].value or ""
        ws_formatado[f"F{linha}"] = f"{v1} - {v2}" if v2 and v2 not in v1 else v1
    
    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(4)

    # Criando mapeamento de valores
    def criar_mapa(coluna_chave, coluna_valor):
        mapa = {}
        for linha in range(4, ws_base.max_row + 1):
            chave = ws_base[f"{coluna_chave}{linha}"].value
            valor = ws_base[f"{coluna_valor}{linha}"].value
            if chave:
                mapa[chave] = valor
        return mapa
    mapa_1 = criar_mapa("K", "C")
    mapa_2 = criar_mapa("C", "D")
    
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"E{linha}"] = mapa_1.get(ws_formatado[f"D{linha}"].value, "Não encontrado")
        ws_formatado[f"F{linha}"] = mapa_2.get(ws_formatado[f"E{linha}"].value, "Não encontrado")

    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)
    ws_formatado.delete_cols(6)


    def copiar_sem_coluna(ws_origem, nome_aba, coluna_excluir):
        ws_novo = wb_principal.create_sheet(title=nome_aba)
        for row in ws_origem.iter_rows():
            nova_linha = []
            for i, cell in enumerate(row, start=1):
                if i != coluna_excluir:  # Ignora a coluna desejada
                    nova_celula = ws_novo.cell(row=cell.row, column=len(nova_linha) + 1, value=cell.value)
                    nova_linha.append(cell.value)

    # Criar abas sem coluna G e M
    copiar_sem_coluna(ws_formatado, "Venda", 6)
    copiar_sem_coluna(ws_formatado, "Estoque", 7)
    primeira_aba = wb_principal.sheetnames[0]  # Nome da primeira aba
    ws_formatado = wb_principal[primeira_aba]  # Acessa a aba
    wb_principal.remove(ws_formatado)  # Remove a aba

    # Cria o diretório "Tratado" dentro da pasta temporária do Windows
    temp_dir = os.path.join(tempfile.gettempdir(), "Tratado")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    # Define o caminho do arquivo tratado dentro da pasta "Tratado"
    temp_path = os.path.join(temp_dir, "arquivo_tratado.xlsx")
    wb_principal.save(temp_path)
    return temp_path

def tratamento_havan_parcial(arquivo):
    global wb_principal
    wb_base = load_workbook(base)
    ws_base = wb_base.active

    wb_principal = load_workbook(arquivo)
    ws_formatado = wb_principal.active

    # Remove quebras de texto e mesclagens
    for row in ws_formatado.iter_rows():
        for cell in row:
            if cell.alignment:
                cell.alignment = Alignment(wrap_text=False, horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical)
    
    for merge_range in list(ws_formatado.merged_cells.ranges):
        ws_formatado.unmerge_cells(str(merge_range))
    ws_formatado.freeze_panes = None
    
    # Preenchimento de células vazias com valores superiores
    for col in range(1, 6):  
        for row in range(8, ws_formatado.max_row + 1):
            cell = ws_formatado.cell(row=row, column=col)
            cell_acima = ws_formatado.cell(row=row-1, column=col)
            if cell.value is None or cell.value == "":
                cell.value = cell_acima.value
    
    # Limpeza e organização de colunas
    ws_formatado.delete_rows(1, 4)
    ws_formatado.insert_cols(7, 2)
    colunas_a_remover = [10,11,11,12,13,13]
    

    for col in colunas_a_remover:
        ws_formatado.delete_cols(col)
    ws_formatado['G3'], ws_formatado['H3'] = 'SKU GAMA', 'DESC GAMA'
    ws_formatado['M3'], ws_formatado['N3'] = 'Estoque Total', 'Venda Total'
    
    # Cálculo do estoque e vendas totais
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"M{linha}"] = (ws_formatado[f'J{linha}'].value or 0) + (ws_formatado[f'L{linha}'].value or 0)
        ws_formatado[f"N{linha}"] = (ws_formatado[f'I{linha}'].value or 0) + (ws_formatado[f'K{linha}'].value or 0)
    
    # Organiza nomes dos produtos
    ws_formatado.insert_cols(7)
    for linha in range(4, ws_formatado.max_row + 1):
        v1 = ws_formatado[f"E{linha}"].value or ""
        v2 = ws_formatado[f"F{linha}"].value or ""
        ws_formatado[f"G{linha}"] = f"{v1} - {v2}" if v2 and v2 not in v1 else v1
    
    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(4)
    ws_formatado.delete_cols(4)

    # Criando mapeamento de valores
    def criar_mapa(coluna_chave, coluna_valor):
        mapa = {}
        for linha in range(4, ws_base.max_row + 1):
            chave = ws_base[f"{coluna_chave}{linha}"].value
            valor = ws_base[f"{coluna_valor}{linha}"].value
            if chave:
                mapa[chave] = valor
        return mapa
    mapa_1 = criar_mapa("K", "C")
    mapa_2 = criar_mapa("C", "D")
    
    for linha in range(4, ws_formatado.max_row + 1):
        ws_formatado[f"E{linha}"] = mapa_1.get(ws_formatado[f"D{linha}"].value, "Não encontrado")
        ws_formatado[f"F{linha}"] = mapa_2.get(ws_formatado[f"E{linha}"].value, "Não encontrado")

    colunas_a_remover = [4, 6, 6, 6, 6]
    for col in colunas_a_remover:
        ws_formatado.delete_cols(col)   


    def copiar_sem_coluna(ws_origem, nome_aba, coluna_excluir):
        ws_novo = wb_principal.create_sheet(title=nome_aba)
        for row in ws_origem.iter_rows():
            nova_linha = []
            for i, cell in enumerate(row, start=1):
                if i != coluna_excluir:  # Ignora a coluna desejada
                    nova_celula = ws_novo.cell(row=cell.row, column=len(nova_linha) + 1, value=cell.value)
                    nova_linha.append(cell.value)

    # Criar abas sem coluna G e M
    copiar_sem_coluna(ws_formatado, "Venda", 6)
    copiar_sem_coluna(ws_formatado, "Estoque", 7)
    primeira_aba = wb_principal.sheetnames[0]  # Nome da primeira aba
    ws_formatado = wb_principal[primeira_aba]  # Acessa a aba
    wb_principal.remove(ws_formatado)  # Remove a aba

    # Cria o diretório "Tratado" dentro da pasta temporária do Windows
    temp_dir = os.path.join(tempfile.gettempdir(), "Tratado")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    # Define o caminho do arquivo tratado dentro da pasta "Tratado"
    temp_path = os.path.join(temp_dir, "arquivo_tratado.xlsx")
    wb_principal.save(temp_path)
    return temp_path